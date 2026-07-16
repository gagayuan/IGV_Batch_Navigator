"""
TSV/CSV/XLSX 智能检测与位置解析模块
不依赖文件后缀，根据内容自动判断分隔符；xlsx 文件通过 openpyxl 读取
"""

import csv
import io
import os
import logging
from typing import Tuple, List, Optional

from i18n import T

logger = logging.getLogger(__name__)


# ── 分隔符检测（TSV/CSV）────────────────────────────

def detect_delimiter(filepath: str, sample_lines: int = 10) -> str:
    """
    读取文件前 sample_lines 行，分别用 tab 和逗号尝试解析，
    返回列数更一致、列数更多的那个分隔符。
    """
    with open(filepath, "r", encoding="utf-8-sig") as f:
        lines = []
        for i, line in enumerate(f):
            if i >= sample_lines:
                break
            stripped = line.rstrip("\n\r")
            if stripped:
                lines.append(stripped)

    if not lines:
        return "\t"

    return _score_delimiters(lines, ["\t", ","])


def _score_delimiters(lines: List[str], candidates: List[str]) -> str:
    best_delim = "\t"
    best_score = -1

    for delim in candidates:
        score = _score_one_delimiter(lines, delim)
        if score > best_score:
            best_score = score
            best_delim = delim

    return best_delim


def _score_one_delimiter(lines: List[str], delim: str) -> int:
    col_counts = []
    for line in lines:
        reader = csv.reader(io.StringIO(line), delimiter=delim)
        try:
            row = next(reader)
        except StopIteration:
            continue
        col_counts.append(len(row))

    if not col_counts:
        return 0

    unique = set(col_counts)
    if len(unique) > 2:
        return 0

    most_common = max(set(col_counts), key=col_counts.count)
    consistency = col_counts.count(most_common) / len(col_counts)
    if consistency < 0.5:
        return 0

    return int(most_common * consistency * 10)


# ── 统一入口 ─────────────────────────────────────────

def parse_positions(filepath: str,
                    locus_col: str,
                    sample_col: str = "") -> Tuple[List[str],
                                                    Optional[List[str]],
                                                    List[List[str]]]:
    """
    解析表格文件中的位置和可选样本信息。
    根据扩展名自动分发到 TSV/CSV 或 XLSX 解析器。

    Args:
        filepath: 表格文件路径
        locus_col: 包含"染色体:位置"格式数据的列名
        sample_col: 样本名列名（空字符串表示不使用）

    Returns:
        (loci, samples, raw_rows)
        loci:      ["chr1:12345", "chr2:67890", ...]
        samples:   样本名列表，无样本列时为 None
        raw_rows:  原始行数据列表
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx",):
        return _parse_xlsx(filepath, locus_col, sample_col)
    else:
        return _parse_delimited(filepath, locus_col, sample_col)


# ── TSV/CSV 解析 ─────────────────────────────────────

def _parse_delimited(filepath: str,
                     locus_col: str,
                     sample_col: str) -> Tuple[List[str],
                                               Optional[List[str]],
                                               List[List[str]]]:
    delim = detect_delimiter(filepath)

    loci = []
    samples: List[str] = []
    raw_rows = []
    locus_idx = None
    sample_idx = None
    use_sample = bool(sample_col)

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delim)

        for row in reader:
            if not row or all(cell.strip() == "" for cell in row):
                continue

            if locus_idx is None:
                try:
                    locus_idx = row.index(locus_col)
                except ValueError:
                    raise ValueError(
                        T('err_col_not_found', col=locus_col, available=row)
                    )
                if use_sample:
                    try:
                        sample_idx = row.index(sample_col)
                    except ValueError:
                        use_sample = False
                continue

            try:
                locus_text = row[locus_idx].strip()
                if not locus_text or ":" not in locus_text:
                    continue
                loci.append(locus_text)
                raw_rows.append(row)
                if use_sample and sample_idx is not None:
                    s = row[sample_idx].strip() if sample_idx < len(row) else ""
                    samples.append(s)
                elif use_sample:
                    samples.append("")
            except IndexError:
                continue

    return loci, (samples if use_sample else None), raw_rows


# ── XLSX 解析 ────────────────────────────────────────

def _parse_xlsx(filepath: str,
                locus_col: str,
                sample_col: str) -> Tuple[List[str],
                                          Optional[List[str]],
                                          List[List[str]]]:
    """解析 xlsx 文件，遍历所有 sheet 并拼接结果。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(T('err_no_openpyxl'))

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    all_loci: List[str] = []
    all_samples: List[str] = []
    all_raw_rows: List[List[str]] = []
    use_sample = bool(sample_col)
    header_found = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        locus_idx = None
        sample_idx = None
        sheet_use_sample = use_sample

        for row in rows:
            if not row or all(
                cell is None or str(cell).strip() == "" for cell in row
            ):
                continue

            # 将 openpyxl 的 cell 值统一转为字符串
            row_list = [
                str(cell).strip() if cell is not None else "" for cell in row
            ]

            if locus_idx is None:
                # 表头行
                try:
                    locus_idx = row_list.index(locus_col)
                    header_found = True
                except ValueError:
                    # 此 sheet 无目标列，跳过
                    logger.info("Sheet '%s': column '%s' not found, skipping.",
                                sheet_name, locus_col)
                    break
                if sheet_use_sample:
                    try:
                        sample_idx = row_list.index(sample_col)
                    except ValueError:
                        sheet_use_sample = False
                continue

            # 数据行
            try:
                locus_text = row_list[locus_idx]
                if not locus_text or ":" not in locus_text:
                    continue
                all_loci.append(locus_text)
                all_raw_rows.append(row_list)
                if sheet_use_sample and sample_idx is not None:
                    s = row_list[sample_idx] if sample_idx < len(row_list) else ""
                    all_samples.append(s)
                elif sheet_use_sample:
                    all_samples.append("")
            except IndexError:
                continue

    wb.close()

    if not header_found:
        raise ValueError(T('err_no_col_in_xlsx', col=locus_col))

    if use_sample and not all_samples:
        all_samples = []

    return all_loci, (all_samples if use_sample else None), all_raw_rows


# ── 批次划分 ─────────────────────────────────────────

def make_position_batches(loci: List[str], batch_size: int) -> List[List[str]]:
    if batch_size <= 0:
        return [loci]
    return [loci[i:i + batch_size] for i in range(0, len(loci), batch_size)]